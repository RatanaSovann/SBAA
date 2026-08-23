"""Export the approved brief: two example frames per asset, annotated and handed off as Excel.

Milestone 6. The workbook shape matches the manual spreadsheet process this tool replaces --
one sheet per brand, one row per approved example, image with a red box around the asset burned
in -- so the handoff needs no relearning on the ML team's side. This is a one-shot artifact, not
an audit record: only the analyst's final picks are written here, nothing about what was
rejected or why (see CLAUDE.md's "What this is").

The box itself is hand-drawn by the analyst at approval time (`app.py` Step 5, via
streamlit-cropper), not detected automatically -- see CLAUDE.md's "Planned pipeline" for why an
automated detector isn't in the loop yet.
"""

from __future__ import annotations

import io
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw

THUMBNAIL_WIDTH = 320
_INVALID_SHEET_CHARS = re.compile(r"[\\/?*\[\]:]")


def burn_box(image_path: str | Path, bbox: dict, color: str = "red", width: int = 4) -> Image.Image:
    """Draw a rectangle (as returned by streamlit-cropper) onto a copy of the frame."""
    image = Image.open(image_path).convert("RGB")
    left, top = bbox["left"], bbox["top"]
    right, bottom = left + bbox["width"], top + bbox["height"]
    draw = ImageDraw.Draw(image)
    draw.rectangle([left, top, right, bottom], outline=color, width=width)
    return image


def _sanitize_sheet_name(name: str, used: set[str]) -> str:
    cleaned = _INVALID_SHEET_CHARS.sub("_", name).strip() or "Brand"
    cleaned = cleaned[:31]
    candidate = cleaned
    suffix = 2
    while candidate.lower() in used:
        candidate = f"{cleaned[:28]}_{suffix}"
        suffix += 1
    used.add(candidate.lower())
    return candidate


def build_brief_workbook(approvals: dict[tuple[str, str], list[dict]]) -> Workbook:
    """Build the export workbook: one sheet per brand, one row per approved example.

    `approvals` maps (brand, asset) -> up to two entries, each with file_path, frame_index,
    timestamp_sec, and bbox (the streamlit-cropper dict: left/top/width/height).
    """
    wb = Workbook()
    wb.remove(wb.active)  # default blank sheet; only write sheets for brands that have approvals

    by_brand: dict[str, list[tuple[str, dict]]] = {}
    for (brand, asset), entries in approvals.items():
        for entry in entries:
            by_brand.setdefault(brand, []).append((asset, entry))

    used_names: set[str] = set()
    for brand in sorted(by_brand):
        ws = wb.create_sheet(_sanitize_sheet_name(brand, used_names))
        ws.append(["Asset", "Frame Number", "Timestamp (s)", "Image"])
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions[get_column_letter(4)].width = 46

        for row_idx, (asset, entry) in enumerate(by_brand[brand], start=2):
            ws.cell(row=row_idx, column=1, value=asset)
            ws.cell(row=row_idx, column=2, value=entry["frame_index"])
            ws.cell(row=row_idx, column=3, value=round(entry["timestamp_sec"], 2))

            annotated = burn_box(entry["file_path"], entry["bbox"])
            scale = THUMBNAIL_WIDTH / annotated.width
            annotated = annotated.resize((THUMBNAIL_WIDTH, round(annotated.height * scale)))
            buf = io.BytesIO()
            annotated.save(buf, format="PNG")
            buf.seek(0)
            xl_img = XLImage(buf)
            ws.row_dimensions[row_idx].height = xl_img.height * 0.75  # px -> points, roughly
            ws.add_image(xl_img, f"D{row_idx}")

    return wb
